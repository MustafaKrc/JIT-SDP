import os
import openpyxl
from openpyxl import load_workbook
from copy import copy
from openpyxl.formatting.rule import Rule  # Use generic Rule to cover all cases

# Specify the folder containing the Excel files
folder_path = "."  # Change this to your folder path
output_file = "merged_excel.xlsx"

# Create a new workbook for output
merged_wb = openpyxl.Workbook()
default_sheet = merged_wb.active
merged_wb.remove(default_sheet)  # Remove the default empty sheet

# List all Excel files in the folder
files = [f for f in os.listdir(folder_path) if f.endswith(".xlsx") and f != output_file]

print("Files to merge:", files)

for file in files:
    file_path = os.path.join(folder_path, file)
    
    # Load the workbook (with formatting)
    wb = load_workbook(file_path)
    
    for sheet in wb.sheetnames:
        source_sheet = wb[sheet]
        new_sheet = merged_wb.create_sheet(title=f"{file.replace('_classification_reports.xlsx', '')}")

        # Copy all cells, including values, styles, and formatting
        for row in source_sheet.iter_rows():
            for cell in row:
                new_cell = new_sheet.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                
                # Copy static formatting styles
                if cell.has_style:
                    new_cell.fill = copy(cell.fill)  # Copies static background color
                    new_cell.font = copy(cell.font)  # Copies font style
                    new_cell.border = copy(cell.border)  # Copies border styles
                    new_cell.alignment = copy(cell.alignment)  # Copies alignment
                    new_cell.number_format = cell.number_format  # Copies number format

        # Copy Conditional Formatting Rules
        for range_string in source_sheet.conditional_formatting._cf_rules:
            for rule in source_sheet.conditional_formatting._cf_rules[range_string]:
                if isinstance(rule, Rule):  # Rule is the base class for all conditional formatting rules
                    new_sheet.conditional_formatting.add(range_string, rule)

print("Saving merged file...")
merged_wb.save(output_file)
print(f"Merged file saved as {output_file}")
